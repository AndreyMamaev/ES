import pandas as pd
from pandas.api.types import is_numeric_dtype
import PySimpleGUI as sg
import numpy as np
import calendar
from dateutil.relativedelta import relativedelta
from datetime import datetime
from typing import List
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from quarter_dates import (
    get_first_day_of_the_quarter,
    get_last_day_of_the_quarter
)

sg.theme('DarkTeal12')


class Window():
    REQUIRED_COLUMNS = {
        'Номер карточки',
        'ТБ/ЦА',
        'Подразделение',
        'Статус КЗОП',
        'Вид события',
        'Потерпевший СБЕР',
        'Дата подачи ЗОП',
        'Дата возбуждения УД',
        'Дата передачи дела в суд первой инстанции',
        'Подозреваемые',
        'Сумма ущерба',
        'Ущерб возмещенный',
        'Потерпевшие'
    }
    DATE_FORMAT = '%d.%m.%Y'
    CORRECT_TB = {
        'ББ': 'Байкальский банк',
        'ВВБ': 'Волго-Вятский банк',
        'ДВБ': 'Дальневосточный банк',
        'МБ': 'Московский банк',
        'ПБ': 'Поволжский банк',
        'СБ': 'Сибирский банк',
        'СЗБ': 'Северо-Западный банк',
        'СРБ': 'Среднерусский банк',
        'УБ': 'Уральский банк',
        'ЦЧБ': 'Центрально-Черноземный банк',
        'ЮЗБ': 'Юго-Западный банк',
    }
    TB = {
        'Байкальский банк': [
            '-',
            'Бурятское ГОСБ №8601',
            'Читинские ГОСБ №8600',
            'Якутское ГОСБ №8603',
        ],
        'Волго-Вятский банк': [
            '-',
            'Владимирское ГОСБ №8611',
            'Кировское ГОСБ №8612',
            'Марий Эл ГОСБ №8614',
            'Мордовское ГОСБ №8589',
            'Пермское ГОСБ №6984',
            'Татарстан ГОСБ №8610',
            'Удмуртское ГОСБ №8618',
            'Чувашское ГОСБ №8613',
        ],
        'Дальневосточный банк': [
            '-',
            'Биробиджанский ГОСБ №4157',
            'Благовещенский ГОСБ №8636',
            'Головное отделение по Хабаровскому краю',
            'Камчатский ГОСБ №8556',
            'Приморский ГОСБ №8635',
            'Северо-Восточный ГОСБ №8645',
            'Чукотское головное отделение',
            'Южно-Сахалинский ГОСБ №8567',
        ],
        'Московский банк': [
            '-',
        ],
        'Поволжский банк': [
            '-',
            'Астраханское ГОСБ №8625',
            'Волгоградское ГОСБ №8621',
            'Оренбургское ГОСБ №8623',
            'Пензенское ГОСБ №8624',
            'Саратовское отделение №8622',
            'Ульяновское ГОСБ №8588',
        ],
        'Сибирский банк': [
            '-',
            'Алтайское ГОСБ №8644',
            'Кемеровского ГОСБ №8615',
            'Красноярское отделение №8646',
            'Новосибирское ГОСБ №8047',
            'Омское ГОСБ №8634',
            'Томское отделение №8616',
        ],
        'Северо-Западный банк': [
            '-',
            'Архангельское отделение №8637',
            'Вологодское отделение №8638',
            'Калининградское отделение №8626',
            'Карельское отделение №8628',
            'Коми отделение №8617',
            'Мурманское отделение №8627',
            'Новгородское отделение №8629',
            'Псковское отделение №8630',
        ],
        'Среднерусский банк': [
            '-',
            'Брянское ГОСБ №8605',
            'Восточное Головное отделение',
            'Западное Головное отделение',
            'Ивановское ГОСБ №8639',
            'Калужское ГОСБ №8608',
            'Костромское ГОСБ №8640',
            'Рязанское ГОСБ №8606',
            'Северное Головное отделение',
            'Смоленское ГОСБ №8609',
            'Тверское ГОСБ №8607',
            'Тульское ГОСБ №8604',
            'Южное Головное отделение',
            'Ярославское ГОСБ №0017',
        ],
        'Уральский банк': [
            '-',
            'Башкирское ГОСБ №8598',
            'Западно-Сибирское ГОСБ №8647',
            'Курганское ГОСБ №8599',
            'Новоуренгойское ГОСБ №8369',
            'Сургутское ГОСБ №5940',
            'Челябинское ГОСБ №8597',
        ],
        'Центрально-Черноземный банк': [
            '-',
            'Белгородское ГОСБ №8592',
            'Курское ГОСБ № №8596',
            'Липецкое ГОСБ №8593',
            'Орловское ГОСБ №8595',
            'Тамбовское ГОСБ №8594',
        ],
        'Юго-Западный банк': [
            '-',
            'Дагестанское ГОСБ №8590',
            'Ингушское ГОСБ №8633',
            'Кабардино-Балкарское ГОСБ №8631',
            'Калмыцкое ГОСБ №8579',
            'Карачаево-Черкесское ГОСБ №8585',
            'Краснодарское ГОСБ №8619',
            'Ростовское ГОСБ №5221',
            'Северо-Осетинское ГОСБ №8632',
            'Ставропольское ГОСБ №5230',
            'Чеченское ГОСБ №8643',
        ],
    }
    TB_REPLACE = {
        'СиБ': 'СБ',
        'ПВБ': 'ПБ'
    }
    INDIVID_KEYS = {
        'kpkuo': 'Кпкуо',
        'ker': 'Кэр',
        'kevu': 'Кэву',
        'keppl': 'Кэппл',
        'kmpk': 'Кмпк',
    }

    def __init__(self):
        self.individ_report = pd.DataFrame(
            columns=pd.MultiIndex.from_tuples([
                ('Кпкуо', 'Передано в суд / Возбуждено УД'),
                ('Кпкуо', 'Включено'),
                ('Кпкуо', 'Возбуждено УД'),
                ('Кпкуо', 'Передано в суд'),
                ('Кэр', 'Установлено лиц / Возбуждено УД'),
                ('Кэр', 'Возбуждено УД'),
                ('Кэр', 'Установлено лиц'),
                ('Кэву', 'Кэву'),
                ('Кэву', 'Сумма ущерба'),
                ('Кэву', 'Сумма возмещения'),
                ('Кэву', 'Доля возмещения суммарная'),
                ('Кэву', 'Количество'),
                ('Кэппл', 'Переданов суд > 365 дней / Не возбужденные вовремя'),
                ('Кэппл', 'Возбуждено УД'),
                ('Кэппл', 'Передано в суд > 365 дней'),
                ('Кэппл', 'Передано в суд <= 365 дней'),
                ('Кэппл', 'Включено > 365 дней'),
                ('Кэппл', 'Включено <= 365 дней'),
                ('Кмпк', 'Возбуждено УД / Подано ЗОП'),
                ('Кмпк', 'Подано ЗОП'),
                ('Кмпк', 'Возбуждено УД'),
            ]),
        )
        self.individ_registry_table = pd.DataFrame(
            columns=[
                'КПКУО_Номер карточки',
                'КПКУО_Дата возбуждения УД',
                'КПУД_Дата прекращения дела',
            ],
        )
        self.legal_registry_tables = {
            'legal_kvu_kur': {
                'name': 'КВУ_КУР',
                'df': pd.DataFrame(
                    columns=[
                        'КВУ_Наименование',
                        'КВУ_Номер карточки',
                        'КВУ_ТБ',
                    ],
                )
            },
            'legal_kvu_kvupp_kpupp': {
                'name': 'КВУ_КВУПП_КПУПП',
                'df': pd.DataFrame(
                    columns=[
                        'КВУ_Наименование',
                        'КВУ_Номер карточки',
                        'КВУ_ТБ',
                        'Подразделение',
                        'КПУПП_Ущерб причиненный',
                        'КВУПП_Ущерб возмещеннный',
                    ],
                )
            },
            'legal_koup_kur': {
                'name': 'КОУП_КУР',
                'df': pd.DataFrame(
                    columns=[
                        'КОУП_Наименование',
                        'КОУП_Номер карточки',
                        'КОУП_ТБ',
                    ],
                )
            },
            'legal_koup_zpp_vpp': {
                'name': 'КОУП_ЗПП_ВПП',
                'df': pd.DataFrame(
                    columns=[
                        'КОУП_Наименование',
                        'КОУП_Номер карточки',
                        'КОУП_ТБ',
                        'Подразделение',
                    ],
                )
            },
        }
        self.legal_report = pd.DataFrame(
            columns=pd.MultiIndex.from_tuples([
                ('Коуп', 'Коуп'),
                ('Коуп', 'Кол-во'),
                ('Коуп', 'Кол-во ВУД'),
                ('Коуп', 'Кол-во искл.'),
                ('Коуп', 'Кол-во доб.'),
                ('Кву', 'Кву'),
                ('Кву', 'Сумма ущерба'),
                ('Кву', 'Доля возмещения суммарная'),
                ('Кву', 'Сумма возмещения'),
                ('Кву', 'Количество'),
                ('Кву', 'Кол-во искл.'),
                ('Кву', 'Кол-во доб.'),
            ]),
        )
        self.individ_exclusion_inclusion = {
            'kpkuo': {'ex': [], 'in': []},
            'ker': {'ex': [], 'in': []},
            'kevu': {'ex': [], 'in': []},
            'keppl': {'ex': [], 'in': []},
            'kmpk': {'ex': [], 'in': []},
        }
        self.__tabs = []
        self.code = ''
        self._excluded_cards = pd.DataFrame()
        self._included_cards = pd.DataFrame()
        self.excluded_cards = pd.DataFrame()
        self.included_cards = pd.DataFrame()
        self.__open_files()
        self.__get_dfs()
        if self.__check_dfs():
            self.__create_window()

    def __open_files(self):
        """
        Вызывает окно выбора файлов.
        Запысывает пути к выбранным файлам в self.files
        """
        try:
            self.files = sg.popup_get_file(
                'Выберите файлы',
                title="Выберите файлы",
                multiple_files=True,
            ).split(';')
        except Exception:
            self.files = []

    def __get_dfs(self):
        """
        Cоздает словарь self.dfs с элементами типа
        Имя файла: Dataframe из файла
        """
        self.dfs = {
            file_name: pd.read_excel(file_name)
            for file_name in self.files
        }

    def __check_dfs(self) -> bool:
        """
        Проверяет значения словаря self.dfs на наличие
        в df обязательных столбцов.
        В случае отсутсвия файлов или хотя бы одного столбца
        вызывает всплывающее окно и прекращает выполнение программы.
        """
        if len(self.dfs) == 0:
            return False
        for file_name, df in self.dfs.items():
            diff = self.REQUIRED_COLUMNS.difference(df.columns.values)
            if diff != set():
                sg.popup(
                    f'В файле {file_name} отсутствуют столбы:\n'
                    f'{chr(10).join(diff)}\n'
                    f'Проверьте файл и повторите попытку.'
                )
                return False
            for column in ('Ущерб возмещенный',):
                if not is_numeric_dtype(df.loc[:, column]):
                    try:
                        df[column] = df[column].apply(
                            # Аналог перевода даты в числовой тип в Excel
                            # 25569 = количество дней между 00.01.1900 и 01.01.1970
                            lambda x: 25569 + pd.Timestamp(x).timestamp() / 86400
                            if type(x) is datetime else float(x)
                        )
                    except Exception:
                        sg.popup(
                            f'В файле {file_name}, в столбце "{column}" '
                            f'присутствуют нечисловые типы данных.'
                        )
                        return False
        return True

    def __create_window(self):
        """
        Создает рабочее окно программы.
        """
        self.layout = [
            [
                sg.TabGroup([[
                    sg.Tab(
                        filename.split('/')[-1],
                        [[sg.Table(
                            values=df.values.tolist(),
                            headings=df.columns.values.tolist(),
                            vertical_scroll_only=False,
                            justification='center',
                            alternating_row_color=sg.theme_button_color()[1],
                            selected_row_colors='white on black',
                            num_rows=8
                        )]]
                    ) for filename, df in self.dfs.items()
                ]])
            ],
        ]
        self.__create_tabs()
        self.window = sg.Window(
            'ES', self.layout,
            resizable=True,
            finalize=True,
            return_keyboard_events=True,
        )
        self.window.Maximize()
        self.__window_loop()

    def __create_tabs(self):
        """
        Создает группу вкладок для расчета
         коэффициентов физическких и юридических лиц.
        """
        self.__create_individ_tab()
        self.__create_legal_tab()
        self.layout += [
            [sg.TabGroup([self.__tabs], expand_x=True, expand_y=True,)]
        ]

    def __create_individ_tab(self):
        """
        Создает вкладку для работы с физическими лицами.
        """
        self.__tabs.append(sg.Tab('Физические лица', [
            [
                sg.Column([
                    [
                        sg.Text(report, size=(10, 1)),
                        sg.Input(
                            key=f'individ_{key}_date_start',
                            disabled=True,
                            size=(10, 1),
                        ),
                        sg.CalendarButton(
                            "Начало периода",
                            close_when_date_chosen=False,
                            target=f'individ_{key}_date_start',
                            format=self.DATE_FORMAT,
                            size=(12, 1),
                        ),
                        sg.Input(
                            key=f'individ_{key}_date_finish',
                            disabled=True,
                            size=(10, 1)
                        ),
                        sg.CalendarButton(
                            "Конец периода",
                            close_when_date_chosen=False,
                            target=f'individ_{key}_date_finish',
                            format=self.DATE_FORMAT,
                            size=(12, 1)
                        ),
                        sg.Button(
                            "Искл/Доб",
                            size=(7, 1),
                            key=f'individ_{key}_exclusion/inclusion'
                        )
                    ] for key, report in self.INDIVID_KEYS.items()
                ], element_justification='left',),
                sg.Column(
                    [
                        [sg.Input(
                            key='individ_auto_date',
                            disabled=True,
                            size=(14, 1),
                            enable_events=True
                        )],
                        [sg.CalendarButton(
                            "Заполнить для даты",
                            close_when_date_chosen=False,
                            target='individ_auto_date',
                            format=self.DATE_FORMAT,
                            size=(12, 2)
                        )]
                    ], element_justification='left',
                ),
                sg.Column([[]], expand_x=True),
                sg.Column([
                    [
                        sg.Input(
                            key='individ_registry_file',
                            visible=False,
                            enable_events=True,
                        ),
                        sg.FileBrowse(
                            'Загрузить реестр',
                            target='individ_registry_file',
                            file_types=(('Excel', '.xlsx'),),
                        )
                    ],
                    [
                        sg.Table(
                            values=[],
                            headings=(
                                self.individ_registry_table.columns.tolist()
                            ),
                            key='individ_registry_table',
                            num_rows=6,
                            col_widths=3 * [28],
                            auto_size_columns=False,
                            expand_x=True,
                            justification='center',
                            alternating_row_color=(
                                sg.theme_button_color()[1]
                            ),
                            selected_row_colors='white on black',
                        )
                    ],
                ], element_justification='right',)
            ],
            [sg.Text('Результаты')],
            [sg.Table(
                values=[],
                headings=(
                    ['ТБ', 'ГОСБ'] +
                    [': '.join(c) for c in self.individ_report.columns]
                ),
                key='individ_report',
                expand_y=True,
                vertical_scroll_only=False,
                col_widths=[20, 20] + len(self.individ_report.columns) * [10],
                auto_size_columns=False,
                justification='center',
                alternating_row_color=sg.theme_button_color()[1],
                selected_row_colors='white on black',
            )],
            [
                sg.Button(
                    'Расчитать',
                    key='individ_calc_report'
                ),
                sg.Text('', expand_x=True),
                sg.Checkbox(
                    'Без указания исключений и добавлений',
                    key='individ_check_exceptions'
                ),
                sg.Checkbox(
                    'Сохранить только с группировкой',
                    key='individ_check_only_group'
                ),
                sg.Input(
                    visible=False,
                    enable_events=True,
                    key='individ_save_path'
                ),
                sg.FileSaveAs(
                    'Сохранить',
                    key='individ_save_button',
                    file_types=(('Excel', '.xlsx'),),
                    disabled=True
                ),
                sg.Input(
                    visible=False,
                    enable_events=True,
                    key='individ_save_abnormal_path'
                ),
                sg.FileSaveAs(
                    'Сохранить аномальные карточки',
                    key='individ_save_abnormal_button',
                    file_types=(('Excel', '.xlsx'),),
                    disabled=True
                )
            ],
        ]))

    def __create_legal_tab(self):
        """
        Создает вкладку для работы с юридическими лицами.
        """
        self.__tabs.append(sg.Tab(
            'Юридические лица',
            [
                [
                    sg.Radio(
                        "Квартальный отчёт",
                        "legal_type",
                        key='legal_quartal',
                        default=True
                    ),
                    sg.Radio(
                        "Промежуточный отчёт",
                        "legal_type",
                        key='legal_interim'
                    )
                ],
                [
                    sg.Text(
                        "Коуп/Кву",
                        size=(10, 1)
                    ),
                    sg.Input(
                        (
                            datetime.now().date() -
                            relativedelta(
                                years=2,
                                months=11,
                                days=(datetime.now().day - 1)
                            )
                        ).strftime(self.DATE_FORMAT),
                        key='legal_koup_kvu_date_start',
                        disabled=True,
                        size=(10, 1),
                    ),
                    sg.CalendarButton(
                        "Начало периода",
                        close_when_date_chosen=False,
                        target='legal_koup_kvu_date_start',
                        format=self.DATE_FORMAT,
                        size=(12, 1),
                    ),
                    sg.Input(
                        datetime(
                            year=datetime.now().year,
                            month=datetime.now().month,
                            day=calendar.monthrange(
                                datetime.now().year,
                                datetime.now().month
                            )[1]
                        ).strftime(self.DATE_FORMAT),
                        key='legal_koup_kvu_date_finish',
                        disabled=True,
                        size=(10, 1)
                    ),
                    sg.CalendarButton(
                        "Конец периода",
                        close_when_date_chosen=False,
                        target='legal_koup_kvu_date_finish',
                        format=self.DATE_FORMAT,
                        size=(12, 1)
                    ),
                    sg.Input(
                        visible=False,
                        key='legal_koup_kvu_files',
                        enable_events=True,
                    ),
                    sg.FilesBrowse(
                        'Загрузить реестры',
                        target='legal_koup_kvu_files',
                    ),
                    sg.Text('Целевой показатель:'),
                    sg.Spin(
                        values=[round(i, 2) for i in np.arange(0, 100, 0.01)],
                        key='legal_target',
                        initial_value=0,
                        size=(5, 1)
                    )
                ],
                [
                    sg.TabGroup([[
                        sg.Tab(
                            reg['name'],
                            [[sg.Table(
                                values=[],
                                headings=reg['df'].columns.tolist(),
                                key=key,
                                expand_x=True,
                                justification='center',
                                alternating_row_color=(
                                    sg.theme_button_color()[1]
                                ),
                                selected_row_colors='white on black',
                                enable_click_events=True,
                            )], [sg.Button(
                                'Добавить запись',
                                key=f'{key}_add',
                                enable_events=True
                            )]]
                        ) for key, reg in self.legal_registry_tables.items()
                    ]], expand_x=True,)
                ],
                [
                    sg.Table(
                        values=[],
                        headings=(
                            ['ТБ', 'ГОСБ'] +
                            [': '.join(c) for c in self.legal_report.columns]
                        ),
                        key='legal_report',
                        expand_y=True,
                        vertical_scroll_only=False,
                        col_widths=(
                            [20, 20] + len(self.legal_report.columns) * [12]
                        ),
                        auto_size_columns=False,
                        justification='center',
                        alternating_row_color=sg.theme_button_color()[1],
                        selected_row_colors='white on black',
                    )
                ],
                [
                    sg.Button('Расчитать', key='legal_calc_report'),
                    sg.Text('', expand_x=True),
                    sg.Checkbox(
                        'Без указания исключений и добавлений',
                        key='legal_check_exceptions'
                    ),
                    sg.Checkbox(
                        'Сохранить только с группировкой',
                        key='legal_check_only_group'
                    ),
                    sg.Input(
                        visible=False,
                        enable_events=True,
                        key='legal_save_path'
                    ),
                    sg.FileSaveAs(
                        'Сохранить',
                        key='legal_save_button',
                        file_types=(('Excel', '.xlsx'),),
                        disabled=True
                    ),
                ]
            ]
        ))

    def __check_individ_events(self, event, values):
        """
        Проверяет событие и значения переменных окна программы
        для расчета коэффициентов физических лиц.
        """
        if event.endswith('exclusion/inclusion'):
            k = event.split('_')[1]
            win = sg.Window(
                f'Таблица включений/исключений для {self.INDIVID_KEYS[k]}',
                [
                    [sg.Column([[sg.Text(name)], [
                        sg.Multiline(
                            '\n'.join([
                                str(n) for n in
                                self.individ_exclusion_inclusion[k][key]
                            ]),
                            enable_events=True,
                            key=f'{key}clusion_list',
                            size=(25, 10),
                            justification='center',
                            horizontal_scroll=False
                        )
                    ]]
                    ) for name, key in (
                        ('Исключения', 'ex'),
                        ('Включения', 'in')
                    )],
                    [
                        sg.Button('Сохранить', key='save'),
                        sg.Button('Отмена', key='cancel')
                    ]
                ]
            )
            while True:
                e, v = win.read()
                if e in (sg.WIN_CLOSED, 'cancel'):
                    break
                elif e == 'save':
                    self.individ_exclusion_inclusion[k]['ex'] = (
                        [
                            int(''.join(s for s in x if s.isdigit()))
                            for x in v['exclusion_list'].split('\n')
                            if x != ''
                        ]
                    )
                    self.individ_exclusion_inclusion[k]['in'] = (
                        [
                            int(''.join(s for s in x if s.isdigit()))
                            for x in v['inclusion_list'].split('\n')
                            if x != ''
                        ]
                    )
                    break
            win.close()
        match event:
            case 'individ_auto_date':
                auto_date = datetime.strptime(
                    values['individ_auto_date'],
                    self.DATE_FORMAT
                )
                self.window['individ_kpkuo_date_start'].update(
                    get_first_day_of_the_quarter(datetime(
                        year=auto_date.year-1,
                        month=auto_date.month,
                        day=1
                    )).strftime(self.DATE_FORMAT)
                )
                self.window['individ_kpkuo_date_finish'].update(
                    get_last_day_of_the_quarter(datetime(
                        year=auto_date.year-1,
                        month=auto_date.month,
                        day=1
                    )).strftime(self.DATE_FORMAT)
                )
                self.window['individ_ker_date_start'].update(
                    get_first_day_of_the_quarter(datetime(
                        year=auto_date.year-1,
                        month=auto_date.month,
                        day=1
                    )).strftime(self.DATE_FORMAT)
                )
                self.window['individ_ker_date_finish'].update(
                    get_last_day_of_the_quarter(datetime(
                        year=auto_date.year-1,
                        month=auto_date.month,
                        day=1
                    )).strftime(self.DATE_FORMAT)
                )
                self.window['individ_kevu_date_start'].update(
                    (datetime(
                        year=auto_date.year-2,
                        month=auto_date.month,
                        day=1
                    ) + relativedelta(months=1)).strftime(self.DATE_FORMAT)
                )
                self.window['individ_kevu_date_finish'].update(
                    datetime(
                        year=auto_date.year,
                        month=auto_date.month,
                        day=calendar.monthrange(
                            auto_date.year,
                            auto_date.month
                        )[1]
                    ).strftime(self.DATE_FORMAT)
                )
                self.window['individ_keppl_date_start'].update(
                    get_first_day_of_the_quarter(datetime(
                        year=auto_date.year-3,
                        month=auto_date.month,
                        day=1
                    )).strftime(self.DATE_FORMAT)
                )
                self.window['individ_keppl_date_finish'].update(
                    (get_first_day_of_the_quarter(datetime(
                        year=auto_date.year-1,
                        month=auto_date.month,
                        day=1
                    )) - relativedelta(days=1)).strftime(self.DATE_FORMAT)
                )
                self.window['individ_kmpk_date_start'].update(
                    datetime(
                        year=2022,
                        month=7,
                        day=1
                    ).strftime(self.DATE_FORMAT)
                )
                self.window['individ_kmpk_date_finish'].update(
                    datetime(
                        year=auto_date.year,
                        month=auto_date.month,
                        day=calendar.monthrange(
                            auto_date.year,
                            auto_date.month
                        )[1]
                    ).strftime(self.DATE_FORMAT)
                )
            case 'individ_registry_file':
                self.individ_registry_table = pd.DataFrame(
                    columns=self.individ_registry_table.columns.tolist()
                )
                self.individ_registry_table = pd.concat([
                    self.individ_registry_table,
                    pd.read_excel(
                        values['individ_registry_file']
                    )
                ], join='inner')
                self.window['individ_registry_table'].update(
                    self.individ_registry_table.values.tolist()
                )
            case 'individ_save_path':
                to_out_report = self.individ_report
                if values['individ_check_only_group']:
                    gosbs_in_groups = {
                        k: sum(list(v.values()), [])
                        for k, v in ES_individ.GROUP_GOSB.items()
                    }
                    for tb, gosbs in gosbs_in_groups.items():
                        if '-' in gosbs:
                            gosbs += ['Аппарат']
                        to_out_report.query(
                            'ТБ != @tb | '
                            'ГОСБ not in @gosbs',
                            inplace=True
                        )
                if values['individ_check_exceptions']:
                    to_out_report = to_out_report.loc[
                        :,
                        [
                            c for c in to_out_report.columns.tolist()
                            if 'включ' not in c[1].lower()
                        ]
                    ]
                with pd.ExcelWriter(values['individ_save_path']) as writer:
                    to_out_report.to_excel(writer, sheet_name='Полный отчёт')
                    to_out_report.loc[
                        :,
                        [
                            ('Кпкуо', 'Передано в суд / Возбуждено УД'),
                            ('Кэр', 'Установлено лиц / Возбуждено УД'),
                            ('Кэву', 'Кэву'),
                            ('Кэппл', 'Переданов суд > 365 дней / Не возбужденные вовремя'),
                            ('Кмпк', 'Возбуждено УД / Подано ЗОП')
                        ]
                    ].to_excel(writer, sheet_name='Сокращенный отчёт')
                    kpkuo_sheet = to_out_report.query('ГОСБ in @self.TB.keys()') \
                        .reset_index(level='ГОСБ') \
                        .loc[:, [
                            ('Кпкуо', 'Возбуждено УД'),
                            ('Кпкуо', 'Передано в суд'),
                            ('Кпкуо', 'Передано в суд / Возбуждено УД')
                        ]] \
                        .rename(
                            columns={'Передано в суд / Возбуждено УД': '%', },
                            level=1
                        ) \
                        .rename(index=dict(
                            t[::-1] for t in self.CORRECT_TB.items()
                        ))
                    kpkuo_sheet.to_excel(writer, sheet_name='Кпкуо')
                self.__format_xlsx(values['individ_save_path'])
            case 'individ_save_abnormal_path':
                self.individ_abnormal.to_excel(
                    values['individ_save_abnormal_path'],
                    index=False
                )
            case 'individ_calc_report':
                if any([
                    values[v] == '' for v in values if str(v).endswith((
                        '_start', '_finish'
                    )) and str(v).startswith('individ_')
                ]):
                    sg.popup('Заполните даты расчета коэффициентов')
                else:
                    self.individ_report = pd.DataFrame()
                    self.individ_abnormal = pd.DataFrame()
                    self.koup_excluded_cards = pd.DataFrame()
                    self.koup_included_cards = pd.DataFrame()
                    self.kvu_excluded_cards = pd.DataFrame()
                    self.kvu_included_cards = pd.DataFrame()
                    for file_name, dfs in self.dfs.items():
                        for tb in sorted(dfs['ТБ/ЦА'].unique().tolist()):
                            if tb in (
                                'Центральный аппарат (ЦА)',
                                'ПЦП ДСЦ ОСЦ Нижний Новгород'
                            ):
                                continue
                            df = dfs[dfs["ТБ/ЦА"] == tb]
                            report = ES_individ(
                                df=df,
                                kpkuo_date_start=datetime.strptime(
                                    values['individ_kpkuo_date_start'],
                                    self.DATE_FORMAT
                                ).date(),
                                kpkuo_date_finish=datetime.strptime(
                                    values['individ_kpkuo_date_finish'],
                                    self.DATE_FORMAT
                                ).date(),
                                ker_date_start=datetime.strptime(
                                    values['individ_ker_date_start'],
                                    self.DATE_FORMAT
                                ).date(),
                                ker_date_finish=datetime.strptime(
                                    values['individ_ker_date_finish'],
                                    self.DATE_FORMAT
                                ).date(),
                                kmpk_date_start=datetime.strptime(
                                    values['individ_kmpk_date_start'],
                                    self.DATE_FORMAT
                                ).date(),
                                kmpk_date_finish=datetime.strptime(
                                    values['individ_kmpk_date_finish'],
                                    self.DATE_FORMAT
                                ).date(),
                                keppl_date_start=datetime.strptime(
                                    values['individ_keppl_date_start'],
                                    self.DATE_FORMAT
                                ).date(),
                                keppl_date_finish=datetime.strptime(
                                    values['individ_keppl_date_finish'],
                                    self.DATE_FORMAT
                                ).date(),
                                kevu_date_start=datetime.strptime(
                                    values['individ_kevu_date_start'],
                                    self.DATE_FORMAT
                                ).date(),
                                kevu_date_finish=datetime.strptime(
                                    values['individ_kevu_date_finish'],
                                    self.DATE_FORMAT
                                ).date(),
                                exclusion_inclusion=(
                                    self.individ_exclusion_inclusion
                                ),
                                registry=self.individ_registry_table,
                            )
                            self.individ_report = pd.concat([
                                self.individ_report, report.report
                            ])
                            self.individ_abnormal = pd.concat([
                                self.individ_abnormal, report.abnormal
                            ])
                            if len(self.individ_abnormal.index) > 0:
                                self.window[
                                    'individ_save_abnormal_button'
                                ].update(disabled=False)
                            else:
                                self.window[
                                    'individ_save_abnormal_button'
                                ].update(disabled=True)
                    # Добавление строки Сбер
                    self.individ_report.loc[('Сбер', 'Сбер'), :] = self.individ_report[
                        self.individ_report.index.get_level_values('ГОСБ').isin(self.TB.keys())
                    ].sum(axis=0)
                    self.individ_report.loc[('Сбер', 'Сбер'), ('Кпкуо', 'Передано в суд / Возбуждено УД')] = round(
                        100 *
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кпкуо', 'Передано в суд')] /
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кпкуо', 'Возбуждено УД')],
                        2
                    )
                    self.individ_report.loc[('Сбер', 'Сбер'), ('Кэр', 'Установлено лиц / Возбуждено УД')] = round(
                        100 *
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэр', 'Установлено лиц')] /
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэр', 'Возбуждено УД')],
                        2
                    )
                    self.individ_report.loc[('Сбер', 'Сбер'), ('Кэппл', 'Переданов суд > 365 дней / Не возбужденные вовремя')] = round(
                        100 *
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэппл', 'Передано в суд > 365 дней')] /
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэппл', 'Возбуждено УД')],
                        2
                    )
                    self.individ_report.loc[('Сбер', 'Сбер'), ('Кэву', 'Кэву')] = round(
                        100 *
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэву', 'Доля возмещения суммарная')] /
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кэву', 'Количество')],
                        2
                    )
                    self.individ_report.loc[('Сбер', 'Сбер'), ('Кмпк', 'Возбуждено УД / Подано ЗОП')] = round(
                        100 *
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кмпк', 'Возбуждено УД')] /
                        self.individ_report.loc[('Сбер', 'Сбер'), ('Кмпк', 'Подано ЗОП')],
                        2
                    )
                    self.window['individ_report'].update(
                        [list(index) + values for index, values in zip(
                            self.individ_report.index.values.tolist(),
                            self.individ_report.values.tolist()
                        )]
                    )
                    self.window['individ_save_button'].update(disabled=False)

    def __check_legal_events(self, event, values):
        """
        Проверяет событие и значения переменных окна программы
        для расчета коэффициентов юридических лиц.
        """
        if event == 'legal_koup_kvu_files':
            registries = values['legal_koup_kvu_files'].split(';')
            registry_names = set(
                map(
                    lambda x: x + '.xlsx',
                    [r['name'] for r in self.legal_registry_tables.values()]
                )
            )
            if registry_names.difference(
                [registry.split('/')[-1] for registry in registries]
            ) != set():
                sg.popup(
                    f'Загрузите 4 файла с названиями\n'
                    f'{chr(10).join(registry_names)}'
                )
                self.window['legal_koup_kvu_files'].update('')
            else:
                for key, reg in self.legal_registry_tables.items():
                    columns = reg['df'].columns.tolist()
                    registry_file = [
                        registry for registry in registries
                        if reg['name'] + '.xlsx' in registry
                    ][0]
                    new_df = pd.read_excel(registry_file)
                    # if 'Подразделение' in reg['df'].columns.values:
                    #     new_df['Подразделение'] = ''
                    for column in new_df.columns.values:
                        new_df.rename(
                            columns={column: column.strip()},
                            inplace=True
                        )
                        if column.endswith('_ТБ'):
                            for i, value in new_df[column].items():
                                value = value.strip()
                                if value in self.TB_REPLACE.keys():
                                    value = self.TB_REPLACE[value]
                                for tb, tb_full in self.CORRECT_TB.items():
                                    if (
                                        tb in value.split() or
                                        tb_full.split()[0] in value.split()
                                    ):
                                        new_df.loc[
                                            i, column
                                        ] = self.CORRECT_TB[tb]
                                        break
                        elif column.endswith('_Номер карточки'):
                            for i, value in new_df[column].items():
                                new_df.loc[i, column] = value.strip(' №')
                        # elif column == 'Подразделение':
                        #     for i, row in new_df.iterrows():
                        #         number = row[]
                                # new_df.loc[i, column] = 
                    self.legal_registry_tables[key]['df'] = new_df[columns]
                    self.window[key].update(
                        self.legal_registry_tables[key]['df'].values.tolist()
                    )
        elif isinstance(event, tuple) and event[1] == '+CLICKED+':
            if (
                event[2][0] not in (None, -1) and
                self.legal_registry_tables[
                    event[0]
                ]['df'].columns[event[2][1]] == 'Подразделение'
            ):
                for key, reg in self.legal_registry_tables.items():
                    if event[0] == key:
                        tb = reg['df'].loc[
                            event[2][0],
                            [
                                column for column in reg['df'].columns
                                if column.endswith('_ТБ')
                            ][0]
                        ]
                        break
                win = sg.Window(
                    'Выберите подразделение',
                    [
                        [
                            sg.Combo(
                                self.TB[tb],
                                key='combo',
                                expand_x=True,
                                readonly=True
                            )
                        ],
                        [
                            sg.Button('OK', key='ok'),
                            sg.Button('Cancel', key='cancel')
                        ]
                    ],
                    size=(400, 70)
                )
                e, v = win.read()
                if e == 'ok':
                    self.legal_registry_tables[key]['df'].loc[
                        event[2][0],
                        'Подразделение'
                    ] = v['combo']
                    self.window[key].update(
                        self.legal_registry_tables[key]['df'].values.tolist()
                    )
                win.close()
        elif event.endswith('_add'):
            key = event.replace('_add', '')
            df = self.legal_registry_tables[key]['df']
            win = sg.Window(
                'Добавить запись',
                [
                    [
                        sg.Text(
                            'Заполните обязательные поля!',
                            visible=False,
                            key='warn'
                        )
                    ],
                    [
                        sg.Column(
                            [
                                [sg.Text(column)],
                                [
                                    sg.Input(
                                        size=(len(column) + 10, 1),
                                    )
                                    if not column.endswith((
                                        '_ТБ',
                                        'Подразделение'
                                    ))
                                    else (
                                        sg.Combo(
                                            list(self.TB.keys()),
                                            enable_events=True,
                                            key='tb',
                                            readonly=True
                                        )
                                        if column != 'Подразделение'
                                        else sg.Combo(
                                            [],
                                            disabled=True,
                                            key='gosb',
                                            size=(len(column) + 10, 1),
                                            readonly=True
                                        )
                                    )
                                ]
                            ]
                        ) for column in df.columns.values
                    ],
                    [
                        sg.Button('OK', key='ok'),
                        sg.Button('Cancel', key='cancel')
                    ]
                ]
            )
            while True:
                e, v = win.read()
                if e in (sg.WIN_CLOSED, 'cancel'):
                    break
                elif e == 'tb':
                    if 'gosb' in v.keys():
                        win['gosb'].update(
                            values=self.TB[v['tb']],
                            disabled=False
                        )
                elif e == 'ok':
                    if '' in list(v.values())[1:]:
                        win['warn'].update(visible=True)
                    else:
                        self.legal_registry_tables[
                            key
                        ]['df'].loc[len(df.index)] = list(v.values())
                        self.window[key].update(
                            self.legal_registry_tables[
                                key
                            ]['df'].values.tolist()
                        )
                        break
            win.close()
        elif event == 'legal_save_path':
            to_out_report = self.legal_report
            if values['legal_check_only_group']:
                gosbs_in_groups = {
                    k: sum(list(v.values()), [])
                    for k, v in ES_legal.GROUP_GOSB.items()
                }
                for tb, gosbs in gosbs_in_groups.items():
                    if '-' in gosbs:
                        gosbs += ['Аппарат']
                    to_out_report.query(
                        'ТБ != @tb | '
                        'ГОСБ not in @gosbs',
                        inplace=True
                    )
            if values['legal_check_exceptions']:
                to_out_report = to_out_report.loc[
                    :,
                    [
                        c for c in to_out_report.columns.tolist()
                        if 'искл' not in c[1].lower() and 'доб' not in c[1].lower()
                    ]
                ]
            with pd.ExcelWriter(values['legal_save_path']) as writer:
                to_out_report.to_excel(writer, sheet_name='Полный отчёт')
                self.koup_excluded_cards.to_excel(writer, sheet_name='КОУП Исключения', index=False)
                self.koup_included_cards.to_excel(writer, sheet_name='КОУП Включения', index=False)
                self.kvu_excluded_cards.to_excel(writer, sheet_name='КВУ Исключения', index=False)
                self.kvu_included_cards.to_excel(writer, sheet_name='КВУ Включения', index=False)
                to_out_report.loc[:, [('Коуп', 'Коуп'), ('Кву', 'Кву')]].to_excel(
                    writer,
                    sheet_name='Сокращенный отчёт'
                )
                koup_sheet = to_out_report.query('ГОСБ in @self.TB.keys()') \
                    .reset_index(level='ГОСБ') \
                    .loc[:, [
                        ('Коуп', 'Кол-во'),
                        ('Коуп', 'Кол-во ВУД'),
                        ('Коуп', 'Коуп')
                    ]] \
                    .rename(
                        columns={
                            'Коуп': '%',
                            'Кол-во ВУД': 'Возбуждено дел',
                            'Кол-во': 'Подано заявлений'
                        },
                        level=1
                    ) \
                    .rename(index=dict(
                        t[::-1] for t in self.CORRECT_TB.items()
                    ))
                koup_sheet.to_excel(writer, sheet_name='Коуп')
            self.__format_xlsx(values['legal_save_path'])
        elif event == 'legal_calc_report':
            if '' in (
                values['legal_koup_kvu_date_start'],
                values['legal_koup_kvu_date_finish']
            ):
                sg.popup('Заполните даты расчета коэффициентов')
            elif any([
                '' in reg['df']['Подразделение'].values
                for reg in self.legal_registry_tables.values()
                if 'Подразделение' in reg['df'].columns.values
            ]):
                sg.popup(
                    'Заполните значения подразделений в таблицах реестров'
                )
            else:
                self.legal_report = pd.DataFrame()
                self.koup_excluded_cards = pd.DataFrame()
                self.koup_included_cards = pd.DataFrame()
                self.kvu_excluded_cards = pd.DataFrame()
                self.kvu_included_cards = pd.DataFrame()
                for file_name, dfs in self.dfs.items():
                    for tb in sorted(dfs['ТБ/ЦА'].unique().tolist()):
                        if tb in (
                            'Центральный аппарат (ЦА)',
                            'ПЦП ДСЦ ОСЦ Нижний Новгород'
                        ):
                            continue
                        df = dfs[dfs["ТБ/ЦА"] == tb]
                        report = ES_legal(
                            df,
                            registries={
                                key: reg['df'] for key, reg
                                in self.legal_registry_tables.items()
                            },
                            interim=values['legal_interim'],
                            target=(
                                float(values['legal_target'])
                                if str(values['legal_target']).replace(".", "").isnumeric()
                                else values['legal_target']
                            ),
                            koup_kvu_date_start=datetime.strptime(
                                values['legal_koup_kvu_date_start'],
                                self.DATE_FORMAT
                            ).date(),
                            koup_kvu_date_finish=datetime.strptime(
                                values['legal_koup_kvu_date_finish'],
                                self.DATE_FORMAT
                            ).date(),
                        )
                        self.legal_report = pd.concat([
                            self.legal_report,
                            report.report
                        ])
                        self.koup_excluded_cards = pd.concat([
                            self.koup_excluded_cards,
                            report.koup_excluded_cards
                        ])
                        self.koup_included_cards = pd.concat([
                            self.koup_included_cards,
                            report.koup_included_cards
                        ])
                        self.kvu_excluded_cards = pd.concat([
                            self.kvu_excluded_cards,
                            report.kvu_excluded_cards
                        ])
                        self.kvu_included_cards = pd.concat([
                            self.kvu_included_cards,
                            report.kvu_included_cards
                        ])
                # Добавление строки Сбер
                self.legal_report.loc[('Сбер', 'Сбер'), :] = self.legal_report[
                    self.legal_report.index.get_level_values('ГОСБ').isin(self.TB.keys())
                ].sum(axis=0)
                self.legal_report.loc[('Сбер', 'Сбер'), ('Коуп', 'Коуп')] = round(
                    100 *
                    self.legal_report.loc[('Сбер', 'Сбер'), ('Коуп', 'Кол-во ВУД')] /
                    self.legal_report.loc[('Сбер', 'Сбер'), ('Коуп', 'Кол-во')],
                    2
                )
                self.legal_report.loc[('Сбер', 'Сбер'), ('Кву', 'Кву')] = round(
                    100 * self.legal_report.loc[('Сбер', 'Сбер'), ('Кву', 'Доля возмещения суммарная')] /
                    self.legal_report.loc[('Сбер', 'Сбер'), ('Кву', 'Количество')], 2
                )
                self.window['legal_report'].update(
                    [list(index) + values for index, values in zip(
                        self.legal_report.index.values.tolist(),
                        self.legal_report.values.tolist()
                    )]
                )
                self.window['legal_save_button'].update(disabled=False)

    def __window_loop(self):
        """
        Рабочий цикл окна программы.
        """
        while True:
            event, values = self.window.read()
            if event == sg.WIN_CLOSED:
                break
            elif (
                isinstance(event, str) and
                event.startswith('individ_')
            ):
                self.__check_individ_events(event, values)
            elif (
                (
                    isinstance(event, tuple) and
                    event[0].startswith('legal_')
                ) or
                event.startswith('legal_')
            ):
                self.__check_legal_events(event, values)
            elif event is not sg.TIMEOUT_KEY:
                if (
                    len(event) == 1 and
                    ord(event) == 13 and
                    self.code.endswith('upupdowndownleftrightleftrightba')
                ):
                    self.code = ''
                    try:
                        screen = self.window.get_screen_size()
                        win = sg.Window(
                            '', [[sg.Image(filename='./image.png')]],
                            no_titlebar=True,
                            keep_on_top=True,
                            location=(screen[0],  screen[1]),
                            transparent_color=sg.theme_background_color(),
                            finalize=True
                        )
                        size = win.size
                        win.move(screen[0] - size[0], screen[1] - size[1])
                        win.read(timeout=500)
                        win.close()
                    except Exception:
                        pass
                self.code += event.lower().split(':')[0]

    def __format_xlsx(self, file: str):
        """
        Форматирует выходной xlsx файл.
        Принимает абсолютный путь к файлу.
        """
        wb = load_workbook(file)
        thin_line = Side(
                border_style="thin",
                color="000000"
            )
        medium_line = Side(
            border_style="medium",
            color="000000"
        )
        for sheet in wb.sheetnames:
            if (
                "исключения" in sheet.lower() or
                "включения" in sheet.lower()
            ):
                continue
            ws = wb[sheet]
            if sheet in ('Кпкуо', 'Коуп'):
                ws.unmerge_cells(
                    start_row=1,
                    start_column=2,
                    end_row=1,
                    end_column=ws.max_column
                )
                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=ws.max_column
                )
                ws["A1"].value = sheet
                ws["A1"].font = Font(
                    name='SB Sans Display',
                    bold=True,
                    size=24,
                    color='008080'
                )
                for i in range(1, ws.max_column + 1):
                    column = get_column_letter(i)
                    ws.column_dimensions[column].width = 12
                    if ws[f"{column}2"].value is None:
                        ws[f"{column}2"].value = ws[f"{column}3"].value
                        ws[f"{column}2"].font = Font(bold=True)
                    for j in range(1, ws.max_row + 1):
                        cell = ws[f"{column}{j}"]
                        if j in (1, 2):
                            ws.row_dimensions[j].height = 33
                            cell.alignment = Alignment(
                                wrapText=True,
                                horizontal='center',
                                vertical='center',
                            )
                        else:
                            cell.alignment = Alignment(
                                horizontal='center',
                            )
                        if (
                            ws[f"{column}2"].value == '%' and
                            j > 2
                        ):
                            cell.fill = PatternFill('solid', fgColor="99FFCC")
                        cell.border = Border(
                            top=thin_line,
                            bottom=thin_line,
                            right=thin_line,
                            left=thin_line,
                        )
                ws.delete_rows(3)
                for mcr in ws.merged_cells:
                    if 3 < mcr.min_row:
                        mcr.shift(row_shift=-1)
                    elif 3 <= mcr.max_row:
                        mcr.shrink(bottom=1)
                continue
            for i in range(1, ws.max_column + 1):
                column = get_column_letter(i)
                if column in ('A', 'B'):
                    ws.column_dimensions[column].width = 30
                else:
                    ws.column_dimensions[column].width = 15
                if ws[f"{column}2"].value is None:
                    ws[f"{column}2"].value = ws[f"{column}3"].value
                    ws[f"{column}2"].font = Font(bold=True)
                for j in range(1, ws.max_row + 1):
                    cell = ws[f"{column}{j}"]
                    if j == 2:
                        cell.alignment = Alignment(
                            wrapText=True,
                            horizontal='center',
                            vertical='center',
                        )
                        cell.border = Border(
                            top=thin_line,
                            bottom=medium_line,
                            right=Side(**cell.border.right.__dict__),
                            left=Side(**cell.border.left.__dict__),
                        )
                    else:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical='center',
                        )
                    if i == ws.max_column:
                        cell.border = Border(
                            top=Side(**cell.border.top.__dict__),
                            bottom=Side(**cell.border.bottom.__dict__),
                            right=medium_line,
                            left=Side(**cell.border.left.__dict__),
                        )
                    if j == ws.max_row:
                        cell.border = Border(
                            top=Side(**cell.border.top.__dict__),
                            bottom=medium_line,
                            right=Side(**cell.border.right.__dict__),
                            left=Side(**cell.border.left.__dict__),
                        )
                    if (
                        '/' in str(ws[f"{column}2"].value) or
                        str(ws[f"{column}1"].value) == str(ws[f"{column}2"].value) and i > 2
                    ):
                        cell.border = Border(
                            top=Side(**cell.border.top.__dict__),
                            bottom=Side(**cell.border.bottom.__dict__),
                            right=Side(**cell.border.right.__dict__),
                            left=medium_line,
                        )
                        if j > 2:
                            cell.fill = PatternFill('solid', fgColor="c5d9f1")
                    if (
                        ws[f"A{j}"].value == ws[f"B{j}"].value and
                        j > 2
                    ):
                        cell.border = Border(
                            top=medium_line,
                            bottom=Side(**cell.border.bottom.__dict__),
                            right=Side(**cell.border.right.__dict__),
                            left=Side(**cell.border.left.__dict__)
                        )
                        if i > 1:
                            cell.fill = PatternFill('solid', fgColor="92d050")
            ws.freeze_panes = 'C3'
            ws.delete_rows(3)
            for mcr in ws.merged_cells:
                if 3 < mcr.min_row:
                    mcr.shift(row_shift=-1)
                elif 3 <= mcr.max_row:
                    mcr.shrink(bottom=1)
        wb.save(file)


class ES_individ():
    CORRECT_TB = {
        'ББ': 'Байкальский банк',
        'ВВБ': 'Волго-Вятский банк',
        'ДВБ': 'Дальневосточный банк',
        'МБ': 'Московский банк',
        'ПБ': 'Поволжский банк',
        'СБ': 'Сибирский банк',
        'СЗБ': 'Северо-Западный банк',
        'СРБ': 'Среднерусский банк',
        'УБ': 'Уральский банк',
        'ЦЧБ': 'Центрально-Черноземный банк',
        'ЮЗБ': 'Юго-Западный банк',
    }
    GROUP_GOSB = {
        'Среднерусский банк': {
            '[Аппарат СРБ]': [
                'Южное Головное отделение',
                'Восточное Головное отделение',
                'Северное Головное отделение',
                'Западное Головное отделение',
                '-'
            ],
        },
        'Юго-Западный банк': {
            '[Аппарат ЮЗБ]': [
                'Дагестанское ГОСБ №8590',
                'Ингушское ГОСБ №8633',
                'Кабардино-Балкарское ГОСБ №8631',
                'Калмыцкое ГОСБ №8579',
                'Карачаево-Черкесское ГОСБ №8585',
                'Чеченское ГОСБ №8643',
                'Северо-Осетинское ГОСБ №8632',
                '-'
            ]
        },
        'Дальневосточный банк': {
            '[Головное отделение по Хабаровскому краю ДВБ]': [
                'Биробиджанский ГОСБ №4157',
                'Чукотское головное отделение',
                'Головное отделение по Хабаровскому краю',
                'Камчатский ГОСБ №8556',
            ]
        }
    }
    TB_LIST = {
        'Байкальский банк': [
            '-',
            'Бурятское ГОСБ №8601',
            'Читинские ГОСБ №8600',
            'Якутское ГОСБ №8603',
        ],
        'Волго-Вятский банк': [
            '-',
            'Владимирское ГОСБ №8611',
            'Кировское ГОСБ №8612',
            'Марий Эл ГОСБ №8614',
            'Мордовское ГОСБ №8589',
            'Пермское ГОСБ №6984',
            'Татарстан ГОСБ №8610',
            'Удмуртское ГОСБ №8618',
            'Чувашское ГОСБ №8613',
        ],
        'Дальневосточный банк': [
            '-',
            'Биробиджанский ГОСБ №4157',
            'Благовещенский ГОСБ №8636',
            'Головное отделение по Хабаровскому краю',
            'Камчатский ГОСБ №8556',
            'Приморский ГОСБ №8635',
            'Северо-Восточный ГОСБ №8645',
            'Чукотское головное отделение',
            'Южно-Сахалинский ГОСБ №8567',
        ],
        'Московский банк': [
            '-',
        ],
        'Поволжский банк': [
            '-',
            'Астраханское ГОСБ №8625',
            'Волгоградское ГОСБ №8621',
            'Оренбургское ГОСБ №8623',
            'Пензенское ГОСБ №8624',
            'Саратовское отделение №8622',
            'Ульяновское ГОСБ №8588',
        ],
        'Сибирский банк': [
            '-',
            'Алтайское ГОСБ №8644',
            'Кемеровского ГОСБ №8615',
            'Красноярское отделение №8646',
            'Новосибирское ГОСБ №8047',
            'Омское ГОСБ №8634',
            'Томское отделение №8616',
        ],
        'Северо-Западный банк': [
            '-',
            'Архангельское отделение №8637',
            'Вологодское отделение №8638',
            'Калининградское отделение №8626',
            'Карельское отделение №8628',
            'Коми отделение №8617',
            'Мурманское отделение №8627',
            'Новгородское отделение №8629',
            'Псковское отделение №8630',
        ],
        'Среднерусский банк': [
            '-',
            'Брянское ГОСБ №8605',
            'Восточное Головное отделение',
            'Западное Головное отделение',
            'Ивановское ГОСБ №8639',
            'Калужское ГОСБ №8608',
            'Костромское ГОСБ №8640',
            'Рязанское ГОСБ №8606',
            'Северное Головное отделение',
            'Смоленское ГОСБ №8609',
            'Тверское ГОСБ №8607',
            'Тульское ГОСБ №8604',
            'Южное Головное отделение',
            'Ярославское ГОСБ №0017',
        ],
        'Уральский банк': [
            '-',
            'Башкирское ГОСБ №8598',
            'Западно-Сибирское ГОСБ №8647',
            'Курганское ГОСБ №8599',
            'Новоуренгойское ГОСБ №8369',
            'Сургутское ГОСБ №5940',
            'Челябинское ГОСБ №8597',
        ],
        'Центрально-Черноземный банк': [
            '-',
            'Белгородское ГОСБ №8592',
            'Курское ГОСБ № №8596',
            'Липецкое ГОСБ №8593',
            'Орловское ГОСБ №8595',
            'Тамбовское ГОСБ №8594',
        ],
        'Юго-Западный банк': [
            '-',
            'Дагестанское ГОСБ №8590',
            'Ингушское ГОСБ №8633',
            'Кабардино-Балкарское ГОСБ №8631',
            'Калмыцкое ГОСБ №8579',
            'Карачаево-Черкесское ГОСБ №8585',
            'Краснодарское ГОСБ №8619',
            'Ростовское ГОСБ №5221',
            'Северо-Осетинское ГОСБ №8632',
            'Ставропольское ГОСБ №5230',
            'Чеченское ГОСБ №8643',
        ],
    }
    GOSB_RENAME = {
        'адыгейское': 'Краснодарское ГОСБ №8619',
        'ямало-ненец': 'Новоуренгойское ГОСБ №8369',
    }

    def __init__(
        self,
        df,
        kpkuo_date_start,
        kpkuo_date_finish,
        ker_date_start,
        ker_date_finish,
        keppl_date_start,
        keppl_date_finish,
        kevu_date_start,
        kevu_date_finish,
        kmpk_date_start,
        kmpk_date_finish,
        exclusion_inclusion,
        registry
    ):
        self.df = df
        # Валидация ГОСБ
        # self.df.loc[
        #     :, 'Подразделение'
        # ] = self.df.loc[:, 'Подразделение'].apply(
        #     lambda x: (
        #         [
        #             gosb for gosb
        #             in self.TB_LIST[self.df['ТБ/ЦА'].unique()[0].split(' (')[0]]
        #             if x.split()[0] == gosb.split()[0] or (
        #                 len(re.findall(r'\d+', x)) > 0 and
        #                 len(re.findall(r'\d+', gosb)) > 0 and
        #                 re.findall(r'\d+', x)[0] == re.findall(r'\d+', gosb)[0]
        #             )
        #         ][0]
        #     )
        # )
        self.registry = registry
        self.exclusion_inclusion = exclusion_inclusion
        self.kpkuo_date_start = kpkuo_date_start
        self.kpkuo_date_finish = kpkuo_date_finish
        self.ker_date_start = ker_date_start
        self.ker_date_finish = ker_date_finish
        self.keppl_date_start = keppl_date_start
        self.keppl_date_finish = keppl_date_finish
        self.kevu_date_start = kevu_date_start
        self.kevu_date_finish = kevu_date_finish
        self.kmpk_date_start = kmpk_date_start
        self.kmpk_date_finish = kmpk_date_finish
        self.TB = [
            v for k, v in self.CORRECT_TB.items()
            if k in self.df['ТБ/ЦА'].unique()[0]
        ][0]
        self.df.loc[:, 'Подразделение'] = self.df.loc[:, 'Подразделение'].apply(
            lambda x: x.strip() if len([value for gosb, value in self.GOSB_RENAME.items() if gosb in x.lower()]) == 0 else
            [value for gosb, value in self.GOSB_RENAME.items() if gosb in x.lower()][0]
        )
        self.GOSB = self.df['Подразделение'].unique().tolist()
        self.GOSB.sort()
        self.report = pd.DataFrame(
            columns=pd.MultiIndex.from_tuples([
                ('Кпкуо', 'Передано в суд / Возбуждено УД'),
                ('Кпкуо', 'Включено'),
                ('Кпкуо', 'Возбуждено УД'),
                ('Кпкуо', 'Передано в суд'),
                ('Кэр', 'Установлено лиц / Возбуждено УД'),
                ('Кэр', 'Возбуждено УД'),
                ('Кэр', 'Установлено лиц'),
                ('Кэву', 'Кэву'),
                ('Кэву', 'Сумма ущерба'),
                ('Кэву', 'Сумма возмещения'),
                ('Кэву', 'Доля возмещения суммарная'),
                ('Кэву', 'Количество'),
                ('Кэппл', 'Переданов суд > 365 дней / Не возбужденные вовремя'),
                ('Кэппл', 'Возбуждено УД'),
                ('Кэппл', 'Передано в суд > 365 дней'),
                ('Кэппл', 'Передано в суд <= 365 дней'),
                ('Кэппл', 'Включено > 365 дней'),
                ('Кэппл', 'Включено <= 365 дней'),
                ('Кмпк', 'Возбуждено УД / Подано ЗОП'),
                ('Кмпк', 'Подано ЗОП'),
                ('Кмпк', 'Возбуждено УД'),
            ]),
            index=pd.MultiIndex.from_arrays([
                [self.TB] *
                (len(self.GOSB) + len(self.GROUP_GOSB.get(self.TB, {})) + 1),
                [self.TB] +
                list(self.GROUP_GOSB.get(self.TB, {}).keys()) +
                self.GOSB
            ], names=('ТБ', 'ГОСБ')),
            data=0,
        )
        self.abnormal = pd.DataFrame(columns=self.df.columns.values)
        self.update()

    def update(self):
        """
        Фильтрует записи DataFrame.
        Для прошедших фильтрацию создает группы
        к которым относится запись и вызывает функции
        для проверки вхождения в коэффициенты.
        """
        self.df.loc[:, "Номер карточки"] = self.df.loc[:, "Номер карточки"].apply(
            lambda x: int(str(x).split('/')[0])
        )
        self.registry.loc[:, "КПКУО_Номер карточки"] = self.registry.loc[:, "КПКУО_Номер карточки"].apply(
            lambda x: int(str(x).split('/')[0])
        )
        for i, row in self.df.iterrows():
            groups = (
                [self.TB, row['Подразделение']] +
                [
                    group for group, gosbs
                    in self.GROUP_GOSB.get(self.TB, {}).items()
                    if row['Подразделение'].strip() in gosbs
                ]
            )
            if (
                row['Статус КЗОП'] == 'Архив' or
                'Мошенничество в корпоративном кредитовании'
                in row['Вид события']
            ):
                continue
            self.check_kevu(row, groups)
            self.check_kmpk(row, groups)
            self.check_keppl(row, groups)
            self.check_kpkuo(row, groups)
            self.check_ker(row, groups)
        self.calc()

    def check_kmpk(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кмпк.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            row['Потерпевший СБЕР'] in (False, 'false') and
            row['Потерпевшие'] != '-' and
            row['Дата подачи ЗОП'] != '-' and
            row['Дата подачи ЗОП'].date() >= self.kmpk_date_start and
            row['Дата подачи ЗОП'].date() <= self.kmpk_date_finish
        ) or row['Номер карточки'] in self.exclusion_inclusion['kmpk']['in']:
            if row['Номер карточки'] not in self.exclusion_inclusion['kmpk']['ex']:
                for group in groups:
                    self.report.loc[
                        (self.TB, group), ('Кмпк', 'Подано ЗОП')
                    ] += 1
                    if row['Дата возбуждения УД'] != '-':
                        self.report.loc[
                            (self.TB, group), ('Кмпк', 'Возбуждено УД')
                        ] += 1

    def check_kpkuo(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кпкуо.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            row['Потерпевший СБЕР'] in (True, 'true') and
            row['Дата возбуждения УД'] != '-' and
            (
                row['Дата возбуждения УД'].date() >=
                self.kpkuo_date_start
            ) and
            (
                row['Дата возбуждения УД'].date() <=
                self.kpkuo_date_finish
            )
        ) or row['Номер карточки'] in self.exclusion_inclusion['kpkuo']['in']:
            if row['Номер карточки'] not in self.exclusion_inclusion['kpkuo']['ex']:
                for group in groups:
                    self.report.loc[
                        (self.TB, group), ('Кпкуо', 'Возбуждено УД')
                    ] += 1
                    if row['Дата передачи дела в суд первой инстанции'] != 'Отсутствует':
                        if (
                            row['Дата передачи дела в суд первой инстанции'].date() <=
                            row['Дата возбуждения УД'].date() + relativedelta(years=1)
                        ):
                            self.report.loc[
                                (self.TB, group), ('Кпкуо', 'Передано в суд')
                            ] += 1
                    elif row['Номер карточки'] in self.registry["КПКУО_Номер карточки"].values:
                        self.report.loc[
                            (self.TB, group),
                            ("Кпкуо", "Включено")
                        ] += 1
                        self.report.loc[
                            (self.TB, group), ('Кпкуо', 'Передано в суд')
                        ] += 1

    def check_ker(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кэр.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            row['Потерпевший СБЕР'] in (True, 'true') and
            row['Дата возбуждения УД'] != '-' and
            (
                row['Дата возбуждения УД'].date() >=
                self.ker_date_start
            ) and
            (
                row['Дата возбуждения УД'].date() <=
                self.ker_date_finish
            )
        ) or row['Номер карточки'] in self.exclusion_inclusion['ker']['in']:
            if row['Номер карточки'] not in self.exclusion_inclusion['ker']['ex']:
                for group in groups:
                    self.report.loc[
                        (self.TB, group), ('Кэр', 'Возбуждено УД')
                    ] += 1
                    if row['Подозреваемые'] in (True, 'true'):
                        self.report.loc[
                            (self.TB, group), ('Кэр', 'Установлено лиц')
                        ] += 1

    def check_keppl(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кэппл.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            row['Потерпевший СБЕР'] in (True, 'true') and
            row['Дата возбуждения УД'] != '-' and
            row['Дата возбуждения УД'].date() >= self.keppl_date_start and
            row['Дата возбуждения УД'].date() <= self.keppl_date_finish
        ) or row['Номер карточки'] in self.exclusion_inclusion['keppl']['in']:
            if row['Номер карточки'] not in self.exclusion_inclusion['keppl']['ex']:
                for group in groups:
                    self.report.loc[(self.TB, group), ('Кэппл', 'Возбуждено УД')] += 1
                    if (
                        row['Дата передачи дела в суд первой инстанции'] !=
                        'Отсутствует'
                    ):
                        if row[
                            'Дата передачи дела в суд первой инстанции'
                        ].date() <= (
                            row['Дата возбуждения УД'].date() +
                            relativedelta(years=1)
                        ):
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Передано в суд <= 365 дней')
                            ] += 1
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Возбуждено УД')
                            ] -= 1
                        else:
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Передано в суд > 365 дней')
                            ] += 1

                    elif row['Номер карточки'] in self.exclusion_inclusion['keppl']['in']:
                        self.report.loc[
                            (self.TB, group),
                            ('Кэппл', 'Передано в суд > 365 дней')
                        ] += 1
                    elif row['Номер карточки'] in self.registry['КПКУО_Номер карточки'].values:
                        registry_row = self.registry[self.registry['КПКУО_Номер карточки'] == row['Номер карточки']].iloc[0]
                        if registry_row[
                            'КПУД_Дата прекращения дела'
                        ].date() <= (
                            registry_row['КПКУО_Дата возбуждения УД'].date() +
                            relativedelta(years=1)
                        ):
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Включено <= 365 дней')
                            ] += 1
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Передано в суд <= 365 дней')
                            ] += 1
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Возбуждено УД')
                            ] -= 1
                        else:
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Включено > 365 дней')
                            ] += 1
                            self.report.loc[
                                (self.TB, group),
                                ('Кэппл', 'Передано в суд > 365 дней')
                            ] += 1

    def check_kevu(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кэву.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            row['Потерпевший СБЕР'] in (True, 'true') and
            row['Дата подачи ЗОП'] != '-' and
            row['Дата подачи ЗОП'].date() >= self.kevu_date_start and
            row['Дата подачи ЗОП'].date() <= self.kevu_date_finish
        ) or row['Номер карточки'] in self.exclusion_inclusion['kevu']['in']:
            if row['Номер карточки'] not in self.exclusion_inclusion['kevu']['ex']:
                if row['Ущерб возмещенный'] == '-':
                    row['Ущерб возмещенный'] = 0
                if row['Сумма ущерба'] == '-':
                    row['Сумма ущерба'] = 0
                if (
                    float(row['Сумма ущерба']) == 0 and
                    float(row['Ущерб возмещенный']) != 0
                ):
                    self.abnormal.loc[len(self.abnormal.index)] = row
                    win = sg.Window(
                        'Аномальная карточка',
                        [
                            [sg.Table(
                                [row.values.tolist()],
                                headings=row.index.values.tolist(),
                                vertical_scroll_only=False,
                                justification='center',
                                alternating_row_color=(
                                    sg.theme_button_color()[1]
                                ),
                                selected_row_colors='white on black',
                                num_rows=3
                            )],
                            [
                                sg.Text('Сумма ущерба:', size=(25, 1)),
                                sg.Input(
                                    default_text=row['Сумма ущерба'],
                                    key='sum_damage',
                                    enable_events=True,
                                    expand_x=True
                                )
                            ],
                            [
                                sg.Text('Ущерб возмещенный:', size=(25, 1)),
                                sg.Input(
                                    default_text=row['Ущерб возмещенный'],
                                    key='comp_damage',
                                    enable_events=True,
                                    expand_x=True
                                )
                            ],
                            [
                                sg.Button('Учитывать', key='allow'),
                                sg.Button('Не учитывать', key='dont_allow')
                            ]
                        ],
                        size=(600, 190)
                    )
                    while True:
                        e, v = win.read()
                        if e in (sg.WIN_CLOSED, 'dont_allow'):
                            break
                        elif e == 'allow':
                            row['Сумма ущерба'] = v['sum_damage']
                            row['Ущерб возмещенный'] = v['comp_damage']
                            break
                    win.close()
                    if e == 'allow':
                        for group in groups:
                            if float(row['Сумма ущерба']) > 0:
                                self.report.loc[
                                    (self.TB, group),
                                    ('Кэву', 'Доля возмещения суммарная')
                                ] += (
                                    float(row['Ущерб возмещенный']) /
                                    float(row['Сумма ущерба'])
                                )
                                self.report.loc[
                                    (self.TB, group),
                                    ('Кэву', 'Сумма ущерба')
                                ] += float(row['Сумма ущерба'])
                                self.report.loc[
                                    (self.TB, group),
                                    ('Кэву', 'Сумма возмещения')
                                ] += float(row['Ущерб возмещенный'])
                            self.report.loc[
                                (self.TB, group),
                                ('Кэву', 'Количество')
                            ] += 1
                elif float(row['Сумма ущерба']) > 0:
                    for group in groups:
                        summ = (
                            float(row['Ущерб возмещенный']) /
                            float(row['Сумма ущерба'])
                        )
                        if summ > 1:
                            summ = 1
                        self.report.loc[
                            (self.TB, group),
                            ('Кэву', 'Доля возмещения суммарная')
                        ] += summ
                        self.report.loc[
                            (self.TB, group),
                            ('Кэву', 'Сумма ущерба')
                        ] += float(row['Сумма ущерба'])
                        self.report.loc[
                            (self.TB, group),
                            ('Кэву', 'Сумма возмещения')
                        ] += float(row['Ущерб возмещенный'])
                        self.report.loc[
                            (self.TB, group),
                            ('Кэву', 'Количество')
                        ] += 1

    def calc(self):
        """
        Расчитывает коэффициенты в результирующем DF.
        """
        self.report[('Кпкуо', 'Передано в суд / Возбуждено УД')] = round(
            100 *
            self.report[('Кпкуо', 'Передано в суд')] /
            self.report[('Кпкуо', 'Возбуждено УД')],
            2
        )
        self.report[('Кэр', 'Установлено лиц / Возбуждено УД')] = round(
            100 *
            self.report[('Кэр', 'Установлено лиц')] /
            self.report[('Кэр', 'Возбуждено УД')],
            2
        )
        self.report[('Кэппл', 'Переданов суд > 365 дней / Не возбужденные вовремя')] = round(
            100 *
            self.report[('Кэппл', 'Передано в суд > 365 дней')] /
            self.report[('Кэппл', 'Возбуждено УД')],
            2
        )
        self.report[('Кэву', 'Кэву')] = round(
            100 *
            self.report[('Кэву', 'Доля возмещения суммарная')] /
            self.report[('Кэву', 'Количество')],
            2
        )
        self.report[('Кмпк', 'Возбуждено УД / Подано ЗОП')] = round(
            100 *
            self.report[('Кмпк', 'Возбуждено УД')] /
            self.report[('Кмпк', 'Подано ЗОП')],
            2
        )
        self.report.replace([np.inf, -np.inf, np.nan], 0, inplace=True)
        self.report.rename(
            index={'-': 'Аппарат'},
            inplace=True,
            errors='ignore'
        )


class ES_legal():
    CORRECT_TB = {
        'ББ': 'Байкальский банк',
        'ВВБ': 'Волго-Вятский банк',
        'ДВБ': 'Дальневосточный банк',
        'МБ': 'Московский банк',
        'ПБ': 'Поволжский банк',
        'СБ': 'Сибирский банк',
        'СЗБ': 'Северо-Западный банк',
        'СРБ': 'Среднерусский банк',
        'УБ': 'Уральский банк',
        'ЦЧБ': 'Центрально-Черноземный банк',
        'ЮЗБ': 'Юго-Западный банк',
    }
    GROUP_GOSB = {
        'Среднерусский банк': {
            '[Аппарат СРБ]': [
                'Южное Головное отделение',
                'Восточное Головное отделение',
                'Северное Головное отделение',
                'Западное Головное отделение',
                '-'
            ],
        },
        'Юго-Западный банк': {
            '[Аппарат ЮЗБ]': [
                'Дагестанское ГОСБ №8590',
                'Ингушское ГОСБ №8633',
                'Кабардино-Балкарское ГОСБ №8631',
                'Калмыцкое ГОСБ №8579',
                'Карачаево-Черкесское ГОСБ №8585',
                'Чеченское ГОСБ №8643',
                'Северо-Осетинское ГОСБ №8632',
                '-'
            ]
        },
        'Дальневосточный банк': {
            '[Головное отделение по Хабаровскому краю ДВБ]': [
                '-',
                'Биробиджанский ГОСБ №4157',
                'Чукотское головное отделение',
                'Головное отделение по Хабаровскому краю',
                'Камчатский ГОСБ №8556',
            ]
        }
    }
    TB_LIST = {
        'Байкальский банк': [
            '-',
            'Бурятское ГОСБ №8601',
            'Читинские ГОСБ №8600',
            'Якутское ГОСБ №8603',
        ],
        'Волго-Вятский банк': [
            '-',
            'Владимирское ГОСБ №8611',
            'Кировское ГОСБ №8612',
            'Марий Эл ГОСБ №8614',
            'Мордовское ГОСБ №8589',
            'Пермское ГОСБ №6984',
            'Татарстан ГОСБ №8610',
            'Удмуртское ГОСБ №8618',
            'Чувашское ГОСБ №8613',
        ],
        'Дальневосточный банк': [
            '-',
            'Биробиджанский ГОСБ №4157',
            'Благовещенский ГОСБ №8636',
            'Головное отделение по Хабаровскому краю',
            'Камчатский ГОСБ №8556',
            'Приморский ГОСБ №8635',
            'Северо-Восточный ГОСБ №8645',
            'Чукотское головное отделение',
            'Южно-Сахалинский ГОСБ №8567',
        ],
        'Московский банк': [
            '-',
        ],
        'Поволжский банк': [
            '-',
            'Астраханское ГОСБ №8625',
            'Волгоградское ГОСБ №8621',
            'Оренбургское ГОСБ №8623',
            'Пензенское ГОСБ №8624',
            'Саратовское отделение №8622',
            'Ульяновское ГОСБ №8588',
        ],
        'Сибирский банк': [
            '-',
            'Алтайское ГОСБ №8644',
            'Кемеровского ГОСБ №8615',
            'Красноярское отделение №8646',
            'Новосибирское ГОСБ №8047',
            'Омское ГОСБ №8634',
            'Томское отделение №8616',
        ],
        'Северо-Западный банк': [
            '-',
            'Архангельское отделение №8637',
            'Вологодское отделение №8638',
            'Калининградское отделение №8626',
            'Карельское отделение №8628',
            'Коми отделение №8617',
            'Мурманское отделение №8627',
            'Новгородское отделение №8629',
            'Псковское отделение №8630',
        ],
        'Среднерусский банк': [
            '-',
            'Брянское ГОСБ №8605',
            'Восточное Головное отделение',
            'Западное Головное отделение',
            'Ивановское ГОСБ №8639',
            'Калужское ГОСБ №8608',
            'Костромское ГОСБ №8640',
            'Рязанское ГОСБ №8606',
            'Северное Головное отделение',
            'Смоленское ГОСБ №8609',
            'Тверское ГОСБ №8607',
            'Тульское ГОСБ №8604',
            'Южное Головное отделение',
            'Ярославское ГОСБ №0017',
        ],
        'Уральский банк': [
            '-',
            'Башкирское ГОСБ №8598',
            'Западно-Сибирское ГОСБ №8647',
            'Курганское ГОСБ №8599',
            'Новоуренгойское ГОСБ №8369',
            'Сургутское ГОСБ №5940',
            'Челябинское ГОСБ №8597',
        ],
        'Центрально-Черноземный банк': [
            '-',
            'Белгородское ГОСБ №8592',
            'Курское ГОСБ № №8596',
            'Липецкое ГОСБ №8593',
            'Орловское ГОСБ №8595',
            'Тамбовское ГОСБ №8594',
        ],
        'Юго-Западный банк': [
            '-',
            'Дагестанское ГОСБ №8590',
            'Ингушское ГОСБ №8633',
            'Кабардино-Балкарское ГОСБ №8631',
            'Калмыцкое ГОСБ №8579',
            'Карачаево-Черкесское ГОСБ №8585',
            'Краснодарское ГОСБ №8619',
            'Ростовское ГОСБ №5221',
            'Северо-Осетинское ГОСБ №8632',
            'Ставропольское ГОСБ №5230',
            'Чеченское ГОСБ №8643',
        ],
    }
    GOSB_RENAME = {
        'адыгейское': 'Краснодарское ГОСБ №8619',
        'ямало-ненец': 'Новоуренгойское ГОСБ №8369',
    }

    def __init__(
        self,
        df,
        interim,
        target,
        koup_kvu_date_start,
        koup_kvu_date_finish,
        registries
    ):
        self.df = df
        # Валидация ГОСБ
        # self.df.loc[
        #     :, 'Подразделение'
        # ] = self.df.loc[:, 'Подразделение'].apply(
        #     lambda x: (
        #         [
        #             gosb for gosb
        #             in self.TB_LIST[self.df['ТБ/ЦА'].unique()[0].split(' (')[0]]
        #             if x.split()[0] == gosb.split()[0] or (
        #                 len(re.findall(r'\d+', x)) > 0 and
        #                 len(re.findall(r'\d+', gosb)) > 0 and
        #                 re.findall(r'\d+', x)[0] == re.findall(r'\d+', gosb)[0]
        #             )
        #         ][0]
        #     )
        # )
        self.interim = interim
        self.target = target
        self.koup_kvu_date_start = koup_kvu_date_start
        self.koup_kvu_date_finish = koup_kvu_date_finish
        self.registries = registries
        self.TB = [
            v for k, v in self.CORRECT_TB.items()
            if k in self.df['ТБ/ЦА'].unique()[0]
        ][0]
        self.koup_excluded_cards = pd.DataFrame(columns=self.df.columns.values)
        self.kvu_excluded_cards = pd.DataFrame(columns=self.df.columns.values)
        self.koup_included_cards = pd.DataFrame(columns=self.registries['legal_koup_zpp_vpp'].columns.values)
        self.kvu_included_cards = pd.DataFrame(columns=self.registries['legal_kvu_kvupp_kpupp'].columns.values)
        self.df.loc[:, 'Подразделение'] = self.df.loc[:, 'Подразделение'].apply(
            lambda x: x.strip() if len([value for gosb, value in self.GOSB_RENAME.items() if gosb in x.lower()]) == 0 else
            [value for gosb, value in self.GOSB_RENAME.items() if gosb in x.lower()][0]
        )
        self.GOSB = df['Подразделение'].unique().tolist()
        self.GOSB.sort()
        self.report = pd.DataFrame(
            columns=pd.MultiIndex.from_tuples([
                ('Коуп', 'Коуп'),
                ('Коуп', 'Кол-во'),
                ('Коуп', 'Кол-во ВУД'),
                ('Коуп', 'Кол-во искл.'),
                ('Коуп', 'Кол-во доб.'),
                ('Кву', 'Кву'),
                ('Кву', 'Сумма ущерба'),
                ('Кву', 'Доля возмещения суммарная'),
                ('Кву', 'Сумма возмещения'),
                ('Кву', 'Количество'),
                ('Кву', 'Кол-во искл.'),
                ('Кву', 'Кол-во доб.'),
            ]),
            index=pd.MultiIndex.from_arrays([
                [self.TB] *
                (len(self.GOSB) + len(self.GROUP_GOSB.get(self.TB, {})) + 1),
                [self.TB] +
                list(self.GROUP_GOSB.get(self.TB, {}).keys()) +
                self.GOSB
            ], names=('ТБ', 'ГОСБ')),
            data=0,
        )
        self.interim_list = pd.DataFrame(columns=self.df.columns)
        self.update()

    def update(self):
        """
        Фильтрует записи DataFrame.
        Для прошедших фильтрацию создает группы
        к которым относится запись и вызывает функции
        для проверки вхождения в коэффициенты.
        """
        self.df.loc[:, "Номер карточки"] = self.df.loc[:, "Номер карточки"].apply(
            lambda x: int(str(x).split('/')[0])
        )
        for reg, name in (
            ('legal_kvu_kvupp_kpupp', 'КВУ'),
            ('legal_koup_zpp_vpp', 'КОУП'),
            ('legal_koup_kur', 'КОУП'),
            ('legal_kvu_kur', 'КВУ'),
        ):
            self.registries[reg][f'{name}_Номер карточки'] = (
                self.registries[reg][f'{name}_Номер карточки'].apply(
                    lambda x: int(str(x).split('/')[0])
                )
            )
        for i, row in self.df.iterrows():
            groups = (
                [self.TB, row['Подразделение']] +
                [
                    group for group, gosbs
                    in self.GROUP_GOSB.get(self.TB, {}).items()
                    if row['Подразделение'].strip() in gosbs
                ]
            )
            if (
                row['Статус КЗОП'] == 'Архив' or
                'Мошенничество в корпоративном кредитовании'
                not in row['Вид события'] or
                row['Дата подачи ЗОП'] == '-' or
                row['Дата подачи ЗОП'].date() < self.koup_kvu_date_start or
                row['Дата подачи ЗОП'].date() > self.koup_kvu_date_finish
            ):
                continue
            self.check_koup(row, groups)
            self.check_kvu(row, groups)
        self.check_kvu_kvupp_kpupp()
        self.check_koup_zpp_vpp()
        self.check_kvu_interim()
        self.calc()

    def check_koup(self, row: pd.Series, groups: List = []):
        """
        Проверка для Коуп.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            self.interim and
            row['Дата подачи ЗОП'].date() >= (
                self.koup_kvu_date_finish -
                relativedelta(months=6)
            ) and
            row['Дата возбуждения УД'] == '-'
        ):
            return
        for group in groups:
            if (
                row['Номер карточки'] in
                self.registries['legal_koup_kur']['КОУП_Номер карточки'].values
            ):
                self.report.loc[(self.TB, group), ('Коуп', 'Кол-во искл.')] += 1
                self.koup_excluded_cards.loc[len(self.koup_excluded_cards.index)] = row
                continue
            self.report.loc[(self.TB, group), ('Коуп', 'Кол-во')] += 1
            if row['Дата возбуждения УД'] != '-':
                self.report.loc[(self.TB, group), ('Коуп', 'Кол-во ВУД')] += 1

    def check_kvu(self, row: pd.Series, groups: List = []):
        """
        Проверка для Кву.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        if (
            self.interim and
            row['Дата подачи ЗОП'].date() >= (
                self.koup_kvu_date_finish -
                relativedelta(months=6)
            )
        ):
            self.interim_list.loc[len(self.interim_list.index)] = row
            return
        for group in groups:
            if (
                row['Номер карточки'] in
                self.registries['legal_kvu_kur']['КВУ_Номер карточки'].values
            ):
                self.report.loc[(self.TB, group), ('Кву', 'Кол-во искл.')] += 1
                self.kvu_excluded_cards.loc[len(self.kvu_excluded_cards.index)] = row
                continue
            if row['Сумма ущерба'] != '-':
                if row['Ущерб возмещенный'] == '-':
                    row['Ущерб возмещенный'] = 0
                self.report.loc[(self.TB, group), ('Кву', 'Доля возмещения суммарная')] += (
                    float(row['Ущерб возмещенный']) /
                    float(row['Сумма ущерба'])
                )
                self.report.loc[(self.TB, group), ('Кву', 'Сумма ущерба')] += float(row['Сумма ущерба'])
                self.report.loc[(self.TB, group), ('Кву', 'Сумма возмещения')] += float(row['Ущерб возмещенный'])
                self.report.loc[(self.TB, group), ('Кву', 'Количество')] += 1

    def check_kvu_interim(self):
        """
        Проверяет отложенные карточки для промежуточного отчета.
        В случае если карточка увеличивает Кву, она учитывается.
        """
        for i, row in self.interim_list.iterrows():
            groups = (
                [self.TB, row['Подразделение']] +
                [
                    group for group, gosbs
                    in self.GROUP_GOSB.get(self.TB, {}).items()
                    if row['Подразделение'].strip() in gosbs
                ]
            )
            for group in groups:
                if (
                    row['Номер карточки'] in
                    self.registries['legal_kvu_kur']['КВУ_Номер карточки'].values
                ):
                    self.report.loc[(self.TB, group), ('Кву', 'Кол-во искл.')] += 1
                    self.kvu_excluded_cards.loc[len(self.kvu_excluded_cards.index)] = row
                    continue
                if row['Сумма ущерба'] != '-':
                    if row['Ущерб возмещенный'] == '-':
                        row['Ущерб возмещенный'] = 0
                    kvu = round(
                        100 * self.report.loc[(self.TB, group), ('Кву', 'Доля возмещения суммарная')] /
                        self.report.loc[(self.TB, group), ('Кву', 'Количество')],
                        2
                    )
                    check_kvu = round(
                        100 * (self.report.loc[(self.TB, group), ('Кву', 'Доля возмещения суммарная')] + (float(row['Ущерб возмещенный']) / float(row['Сумма ущерба']))) /
                        (self.report.loc[(self.TB, group), ('Кву', 'Количество')] + 1),
                        2
                    )
                    if check_kvu >= kvu:
                        self.report.loc[(self.TB, group), ('Кву', 'Доля возмещения суммарная')] += (
                            float(row['Ущерб возмещенный']) /
                            float(row['Сумма ущерба'])
                        )
                        self.report.loc[(self.TB, group), ('Кву', 'Сумма ущерба')] += float(row['Сумма ущерба'])
                        self.report.loc[(self.TB, group), ('Кву', 'Сумма возмещения')] += float(row['Ущерб возмещенный'])
                        self.report.loc[(self.TB, group), ('Кву', 'Количество')] += 1

    def check_kvu_kvupp_kpupp(self):
        """
        Проверка реестра для добавлений в Кву.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        for i, row in self.registries['legal_kvu_kvupp_kpupp'].iterrows():
            if row['КВУ_ТБ'] == self.TB:
                groups = (
                    [self.TB, row['Подразделение']] +
                    [
                        group for group, gosbs
                        in self.GROUP_GOSB.get(self.TB, {}).items()
                        if row['Подразделение'].strip() in gosbs
                    ]
                )
                if row['КВУ_Номер карточки'] in self.df['Номер карточки']:
                    continue
                for group in groups:
                    self.report.loc[(self.TB, group), ('Кву', 'Доля возмещения суммарная')] += (
                        row['КВУПП_Ущерб возмещеннный'] /
                        row['КПУПП_Ущерб причиненный']
                    )
                    self.report.loc[(self.TB, group), ('Кву', 'Сумма ущерба')] += row['КПУПП_Ущерб причиненный']
                    self.report.loc[
                        (self.TB, group), ('Кву', 'Сумма возмещения')
                    ] += row['КВУПП_Ущерб возмещеннный']
                    self.report.loc[
                        (self.TB, group), ('Кву', 'Количество')
                    ] += 1
                    self.report.loc[
                        (self.TB, group), ('Кву', 'Кол-во доб.')
                    ] += 1
                    self.kvu_included_cards.loc[len(self.kvu_included_cards.index)] = row

    def check_koup_zpp_vpp(self):
        """
        Проверка реестра для добавлений в Коуп.
        Заполняет поля в результирующем DF
        в случае соответствия условиям.
        """
        for i, row in self.registries['legal_koup_zpp_vpp'].iterrows():
            if row['КОУП_ТБ'] == self.TB:
                groups = (
                    [self.TB, row['Подразделение']] +
                    [
                        group for group, gosbs
                        in self.GROUP_GOSB.get(self.TB, {}).items()
                        if row['Подразделение'].strip() in gosbs
                    ]
                )
                if row['КОУП_Номер карточки'] in self.df['Номер карточки']:
                    continue
                for group in groups:
                    self.report.loc[(self.TB, group), ('Коуп', 'Кол-во')] += 1
                    self.report.loc[(self.TB, group), ('Коуп', 'Кол-во доб.')] += 1
                    self.report.loc[(self.TB, group), ('Коуп', 'Кол-во ВУД')] += 1
                    self.koup_included_cards.loc[len(self.koup_included_cards.index)] = row

    def calc(self):
        """
        Расчитывает коэффициенты в результирующем DF.
        """
        self.report[('Коуп', 'Коуп')] = round(
            100 *
            self.report[('Коуп', 'Кол-во ВУД')] /
            self.report[('Коуп', 'Кол-во')],
            2
        )
        self.report[('Кву', 'Кву')] = round(
            100 *
            self.report[('Кву', 'Доля возмещения суммарная')] /
            self.report[('Кву', 'Количество')],
            2
        )
        self.report.replace([np.inf, -np.inf], 0, inplace=True)
        self.report[('Коуп', 'Коуп')].replace(
            np.nan,
            self.target,
            inplace=True
        )
        self.report[('Кву', 'Кву')].replace(
            np.nan,
            '-',
            inplace=True
        )
        self.report.rename(
            index={'-': 'Аппарат'},
            inplace=True,
            errors='ignore'
        )
        self.koup_excluded_cards.drop_duplicates(inplace=True)
        self.kvu_excluded_cards.drop_duplicates(inplace=True)
        self.koup_included_cards.drop_duplicates(inplace=True)
        self.kvu_included_cards.drop_duplicates(inplace=True)


if __name__ == '__main__':
    Window()
